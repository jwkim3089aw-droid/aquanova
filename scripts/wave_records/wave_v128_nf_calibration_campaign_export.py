from __future__ import annotations

import json
from collections import Counter
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
WAVE_DIR = ROOT / "scripts/wave_records"

SOURCE_XLSX = (
    WAVE_DIR
    / "AquaNova_WAVE_RO_NF_Experimental_Test_Matrix_V46_2026-07-03.xlsx"
)

OUTPUT_XLSX = (
    WAVE_DIR
    / "AquaNova_WAVE_NF_Calibration_Matrix_V128_2026-07-16.xlsx"
)

PLAN_TEMPLATE = (
    WAVE_DIR
    / "AquaNova_WAVE_V70_production_plan_01_ro_nf_orders_1_3.json"
)

OUTPUT_PLAN = (
    WAVE_DIR
    / "AquaNova_WAVE_V128_nf_calibration_12.json"
)

SOURCE_SHEET = "09_NF_BASELINE"
TARGET_SHEET = "11_NF_CALIBRATION_V128"


REQUIRED_COLUMNS = {
    "Batch_Order",
    "Batch_Group",
    "Run_Enabled",
    "Case_ID",
    "Recommended_PDF_Name",
    "Fresh_Project_Required",
    "Test_Lane",
    "Expected_Tier",
    "Test_Purpose",
    "Run_Status",
    "Notes",
    "WAVE_Library_Selection",
    "Feed_Flow_m3h",
    "Temperature_Min_C",
    "Temperature_Design_C",
    "Temperature_Max_C",
    "Pass_Count",
    "Pass1_Recovery_pct",
    "Pass1_Temperature_Mode",
    "Pass1_Temperature_C",
    "Pass1_Flow_Factor",
    "Pass1_Permeate_Back_Pressure_bar",
    "Pass1_Stage_Count",
    "P1S1_PV",
    "P1S1_Elements_per_PV",
    "P1S1_Membrane",
    "P1S1_Stage_Back_Pressure_bar",
    "P1S1_Boost_Pressure_bar",
    "P1S1_Flow_Factor",
}


MEMBRANES = [
    ("NF270", "NF270-400/34"),
    ("NF90", "NF90-400/34"),
]

FEED_PROFILES = [
    ("LowHardness", "Well Water - Low Hardness"),
    ("MedHardness", "Well Water - Med Hardness"),
    ("HighHardness", "Well Water - High Hardness"),
]

RECOVERIES = [65, 85]


def fail(message: str) -> None:
    print(f"\nFAIL: {message}")
    raise SystemExit(1)


def build_cases() -> list[dict]:
    cases = []
    order = 1

    for membrane_code, membrane_name in MEMBRANES:
        for hardness_code, library_selection in FEED_PROFILES:
            for recovery in RECOVERIES:
                case_id = f"V128_NF_{order:03d}"

                pdf_name = (
                    f"{case_id}_"
                    f"{membrane_code}_"
                    f"{hardness_code}_"
                    f"R{recovery}.pdf"
                )

                cases.append(
                    {
                        "Batch_Order": order,
                        "Batch_Group": "NF_CALIBRATION_V128",
                        "Run_Enabled": "Y",
                        "Case_ID": case_id,
                        "Recommended_PDF_Name": pdf_name,
                        "Fresh_Project_Required": "Y",
                        "Test_Lane": "nf/shared-ro-ui",
                        "Expected_Tier": "experimental",
                        "Test_Purpose": (
                            "NF pressure and product-TDS calibration across "
                            "feed hardness and recovery"
                        ),
                        "Run_Status": "Not Started",
                        "Notes": (
                            "Fresh project per case. Run with "
                            "--allow-experimental-ro. Keep topology fixed."
                        ),
                        "WAVE_Library_Selection": library_selection,
                        "Feed_Flow_m3h": 100,
                        "Temperature_Min_C": 10,
                        "Temperature_Design_C": 25,
                        "Temperature_Max_C": 35,
                        "Pass_Count": 1,
                        "Pass1_Recovery_pct": recovery,
                        "Pass1_Temperature_Mode": "Design",
                        "Pass1_Temperature_C": 25,
                        "Pass1_Flow_Factor": 0.85,
                        "Pass1_Permeate_Back_Pressure_bar": 0,
                        "Pass1_Stage_Count": 1,
                        "P1S1_PV": 10,
                        "P1S1_Elements_per_PV": 6,
                        "P1S1_Membrane": membrane_name,
                        "P1S1_Stage_Back_Pressure_bar": 0,
                        "P1S1_Boost_Pressure_bar": 0,
                        "P1S1_Flow_Factor": 0.85,
                    }
                )

                order += 1

    return cases


def export_workbook(cases: list[dict]) -> None:
    if not SOURCE_XLSX.exists():
        fail(f"source workbook not found: {SOURCE_XLSX}")

    workbook = load_workbook(SOURCE_XLSX)

    if SOURCE_SHEET not in workbook.sheetnames:
        fail(f"source sheet not found: {SOURCE_SHEET}")

    if TARGET_SHEET in workbook.sheetnames:
        del workbook[TARGET_SHEET]

    source_sheet = workbook[SOURCE_SHEET]
    target_sheet = workbook.copy_worksheet(source_sheet)
    target_sheet.title = TARGET_SHEET

    headers = [
        target_sheet.cell(row=1, column=column).value
        for column in range(1, target_sheet.max_column + 1)
    ]

    header_map = {
        str(header): index
        for index, header in enumerate(headers, start=1)
        if header not in (None, "")
    }

    missing_columns = sorted(REQUIRED_COLUMNS - set(header_map))

    if missing_columns:
        fail(f"required columns missing: {missing_columns}")

    # 기존 2행의 서식을 새 캠페인 행에 재사용한다.
    style_template = [
        copy(target_sheet.cell(row=2, column=column)._style)
        for column in range(1, target_sheet.max_column + 1)
    ]

    alignment_template = [
        copy(target_sheet.cell(row=2, column=column).alignment)
        for column in range(1, target_sheet.max_column + 1)
    ]

    fill_template = [
        copy(target_sheet.cell(row=2, column=column).fill)
        for column in range(1, target_sheet.max_column + 1)
    ]

    font_template = [
        copy(target_sheet.cell(row=2, column=column).font)
        for column in range(1, target_sheet.max_column + 1)
    ]

    border_template = [
        copy(target_sheet.cell(row=2, column=column).border)
        for column in range(1, target_sheet.max_column + 1)
    ]

    row_height = target_sheet.row_dimensions[2].height

    if target_sheet.max_row > 1:
        target_sheet.delete_rows(2, target_sheet.max_row - 1)

    for row_number, case in enumerate(cases, start=2):
        for column_number in range(1, len(headers) + 1):
            cell = target_sheet.cell(
                row=row_number,
                column=column_number,
            )

            cell._style = copy(style_template[column_number - 1])
            cell.alignment = copy(alignment_template[column_number - 1])
            cell.fill = copy(fill_template[column_number - 1])
            cell.font = copy(font_template[column_number - 1])
            cell.border = copy(border_template[column_number - 1])

        for key, value in case.items():
            target_sheet.cell(
                row=row_number,
                column=header_map[key],
            ).value = value

        if row_height is not None:
            target_sheet.row_dimensions[row_number].height = row_height

    target_sheet.freeze_panes = "A2"

    end_column = get_column_letter(target_sheet.max_column)
    target_sheet.auto_filter.ref = (
        f"A1:{end_column}{target_sheet.max_row}"
    )

    workbook.save(OUTPUT_XLSX)


def export_plan() -> None:
    if not PLAN_TEMPLATE.exists():
        fail(f"plan template not found: {PLAN_TEMPLATE}")

    plan = json.loads(
        PLAN_TEMPLATE.read_text(encoding="utf-8-sig")
    )

    if not isinstance(plan, dict):
        fail("production plan root must be an object")

    if not isinstance(plan.get("cases"), list):
        fail("production plan template has no cases list")

    plan["cases"] = [
        {
            "id": "NF_CALIBRATION_V128_ORDERS_1_12",
            "kind": "ro_excel",
            "path": OUTPUT_XLSX.name,
            "sheet": TARGET_SHEET,
            "start_order": 1,
            "end_order": 12,
            "allow_experimental_ro": True,
        }
    ]

    replacements = {
        "id": "AquaNova_WAVE_V128_NF_CALIBRATION",
        "name": "AquaNova WAVE V128 NF Calibration",
        "description": (
            "Twelve-case NF270/NF90 calibration campaign across "
            "feed hardness and recovery"
        ),
    }

    for key, value in replacements.items():
        if key in plan:
            plan[key] = value

    OUTPUT_PLAN.write_text(
        json.dumps(plan, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def validate(cases: list[dict]) -> None:
    workbook = load_workbook(
        OUTPUT_XLSX,
        read_only=False,
        data_only=False,
    )

    if TARGET_SHEET not in workbook.sheetnames:
        fail("generated NF calibration sheet is missing")

    sheet = workbook[TARGET_SHEET]

    headers = [
        sheet.cell(row=1, column=column).value
        for column in range(1, sheet.max_column + 1)
    ]

    header_map = {
        str(header): index
        for index, header in enumerate(headers, start=1)
        if header not in (None, "")
    }

    generated_rows = []

    for row_number in range(2, sheet.max_row + 1):
        order = sheet.cell(
            row=row_number,
            column=header_map["Batch_Order"],
        ).value

        if order in (None, ""):
            continue

        generated_rows.append(
            {
                key: sheet.cell(
                    row=row_number,
                    column=header_map[key],
                ).value
                for key in (
                    "Batch_Order",
                    "Case_ID",
                    "WAVE_Library_Selection",
                    "Pass1_Recovery_pct",
                    "P1S1_Membrane",
                    "Fresh_Project_Required",
                )
            }
        )

    if len(generated_rows) != 12:
        fail(
            "expected 12 generated rows, "
            f"actual={len(generated_rows)}"
        )

    case_ids = [row["Case_ID"] for row in generated_rows]

    if len(case_ids) != len(set(case_ids)):
        fail("duplicate Case_ID detected")

    membrane_counts = Counter(
        row["P1S1_Membrane"] for row in generated_rows
    )

    recovery_counts = Counter(
        row["Pass1_Recovery_pct"] for row in generated_rows
    )

    feed_counts = Counter(
        row["WAVE_Library_Selection"] for row in generated_rows
    )

    expected_membranes = {
        "NF270-400/34": 6,
        "NF90-400/34": 6,
    }

    expected_recoveries = {
        65: 6,
        85: 6,
    }

    expected_feeds = {
        "Well Water - Low Hardness": 4,
        "Well Water - Med Hardness": 4,
        "Well Water - High Hardness": 4,
    }

    if dict(membrane_counts) != expected_membranes:
        fail(f"membrane distribution mismatch: {membrane_counts}")

    if dict(recovery_counts) != expected_recoveries:
        fail(f"recovery distribution mismatch: {recovery_counts}")

    if dict(feed_counts) != expected_feeds:
        fail(f"feed distribution mismatch: {feed_counts}")

    if any(
        row["Fresh_Project_Required"] != "Y"
        for row in generated_rows
    ):
        fail("every calibration case must require a fresh project")

    plan = json.loads(
        OUTPUT_PLAN.read_text(encoding="utf-8")
    )

    plan_case = plan["cases"][0]

    if plan_case["path"] != OUTPUT_XLSX.name:
        fail("plan workbook path mismatch")

    if plan_case["sheet"] != TARGET_SHEET:
        fail("plan sheet mismatch")

    if plan_case["start_order"] != 1:
        fail("plan start_order mismatch")

    if plan_case["end_order"] != 12:
        fail("plan end_order mismatch")

    print("=" * 80)
    print("V128 NF CALIBRATION CAMPAIGN EXPORT")
    print("=" * 80)
    print(f"xlsx={OUTPUT_XLSX}")
    print(f"sheet={TARGET_SHEET}")
    print(f"case_count={len(generated_rows)}")
    print(f"membrane_counts={dict(membrane_counts)}")
    print(f"recovery_counts={dict(recovery_counts)}")
    print(f"feed_counts={dict(feed_counts)}")
    print(f"plan={OUTPUT_PLAN}")
    print("\nCASES")

    for case in cases:
        print(
            f"{case['Batch_Order']:02d}. "
            f"{case['Case_ID']} | "
            f"{case['P1S1_Membrane']} | "
            f"{case['WAVE_Library_Selection']} | "
            f"R{case['Pass1_Recovery_pct']}"
        )

    print("\nV128 NF calibration campaign export PASS")


def main() -> int:
    cases = build_cases()

    if len(cases) != 12:
        fail(f"campaign matrix must contain 12 cases: {len(cases)}")

    export_workbook(cases)
    export_plan()
    validate(cases)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
