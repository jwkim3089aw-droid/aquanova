from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
WAVE_DIR = ROOT / "scripts/wave_records"
OUT = WAVE_DIR / "results/_report_corpus/v127_nf_campaign_schema_probe.json"


def scalar_text(obj: dict[str, Any]) -> str:
    values = []

    for key, value in obj.items():
        if not isinstance(value, (dict, list)):
            values.append(f"{key}={value}")

    return " | ".join(values)


def walk(obj: Any, path: str = "$"):
    if isinstance(obj, dict):
        yield path, obj

        for key, value in obj.items():
            yield from walk(value, f"{path}.{key}")

    elif isinstance(obj, list):
        for index, value in enumerate(obj):
            yield from walk(value, f"{path}[{index}]")


def inspect_plans():
    matches = []
    seen = set()

    for plan_path in sorted(WAVE_DIR.glob("AquaNova_WAVE_*.json")):
        try:
            payload = json.loads(plan_path.read_text(encoding="utf-8-sig"))
        except Exception:
            continue

        for object_path, obj in walk(payload):
            text = scalar_text(obj).lower()

            if not any(token in text for token in ("nf270", "nf90", "v46_nf")):
                continue

            signature = json.dumps(obj, sort_keys=True, ensure_ascii=False)

            if signature in seen:
                continue

            seen.add(signature)

            matches.append(
                {
                    "plan_file": str(plan_path),
                    "object_path": object_path,
                    "keys": sorted(obj.keys()),
                    "object": obj,
                }
            )

    return matches


def inspect_excel():
    workbooks = []

    patterns = [
        "*NF*.xlsx",
        "*RO_NF*.xlsx",
        "*Experimental_Test_Matrix_V46*.xlsx",
    ]

    paths = set()

    for pattern in patterns:
        paths.update(WAVE_DIR.glob(pattern))

    for workbook_path in sorted(paths):
        try:
            workbook = load_workbook(
                workbook_path,
                read_only=False,
                data_only=False,
            )
        except Exception as exc:
            workbooks.append(
                {
                    "file": str(workbook_path),
                    "error": str(exc),
                }
            )
            continue

        sheets = []

        for sheet in workbook.worksheets:
            preview = []

            max_row = int(sheet.max_row or 0)
            max_column = int(sheet.max_column or 0)

            if max_row > 0 and max_column > 0:
                for row in sheet.iter_rows(
                    min_row=1,
                    max_row=min(max_row, 8),
                    max_col=max_column,
                    values_only=True,
                ):
                    preview.append(list(row))

            sheets.append(
                {
                    "sheet": sheet.title,
                    "max_row": max_row,
                    "max_column": max_column,
                    "preview": preview,
                }
            )

        workbooks.append(
            {
                "file": str(workbook_path),
                "sheets": sheets,
            }
        )

    return workbooks


def main() -> int:
    plan_matches = inspect_plans()
    workbooks = inspect_excel()

    report = {
        "schema_version": "aquanova.nf_campaign_schema_probe.v127",
        "plan_match_count": len(plan_matches),
        "plan_matches": plan_matches,
        "workbook_count": len(workbooks),
        "workbooks": workbooks,
    }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(
        json.dumps(report, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    print("=" * 80)
    print("V127 NF CAMPAIGN SCHEMA PROBE")
    print("=" * 80)
    print(f"plan_match_count={len(plan_matches)}")
    print(f"workbook_count={len(workbooks)}")
    print(f"report={OUT}")

    print("\nPLAN TEMPLATES")

    for index, item in enumerate(plan_matches[:10], start=1):
        print(f"\n[{index}] {Path(item['plan_file']).name}")
        print(f"path={item['object_path']}")
        print(f"keys={item['keys']}")
        print(
            json.dumps(
                item["object"],
                ensure_ascii=False,
                indent=2,
            )
        )

    print("\nEXCEL WORKBOOKS")

    for workbook in workbooks:
        print(f"\nfile={Path(workbook['file']).name}")

        if workbook.get("error"):
            print(f"error={workbook['error']}")
            continue

        for sheet in workbook["sheets"]:
            print(
                f"sheet={sheet['sheet']} "
                f"rows={sheet['max_row']} "
                f"columns={sheet['max_column']}"
            )

            for row in sheet["preview"]:
                print(row)

    print("\nV127 NF campaign schema probe PASS")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())


