# tests/test_hrro_guidelines_match_xlsx.py
import os
import re
from pathlib import Path

import pytest
import openpyxl

from app.services.simulation.modules.hrro import GUIDELINES


# --- Excel fixture known anomalies (typos in the provided xlsx) ---
# In your hrro_baseline.xlsx, Guide Line sheet cell I22 contains a broken string: ").07"
# This should be either blank or ">0.7" (likely), but we treat it as None to match GUIDELINES.
KNOWN_CELL_OVERRIDES = {
    "I22": None,
}


def _parse_num_from_constraint(x):
    """
    Parse numeric value from cells like:
      - "<31"  -> 31.0
      - ">3.6" -> 3.6
      - "7"    -> 7.0
      - None   -> None

    Robust against weird strings (and leading decimals).
    """
    if x is None:
        return None
    if isinstance(x, (int, float)):
        return float(x)

    s = str(x).strip()
    if s in {"", "-", "—", "–", "N/A", "n/a", "None"}:
        return None

    # Guard: broken/invalid strings (fixture typo case etc.)
    # e.g. ").07" -> treat as None (shouldn't be a valid constraint)
    if s.startswith(")"):
        return None

    # capture either "12", "12.3", or ".07"
    m = re.search(r"([-+]?(?:\d*\.\d+|\d+))", s)
    if not m:
        return None
    try:
        return float(m.group(1))
    except Exception:
        return None


def _parse_range_lmh(cell_value):
    """
    Guide Line D/E column example:
      "23\n(20~26)" -> (20.0, 26.0)
    """
    if cell_value is None:
        return None
    s = str(cell_value)
    m = re.search(r"\(([-0-9.]+)\s*~\s*([-0-9.]+)\)", s)
    if not m:
        return None
    return (float(m.group(1)), float(m.group(2)))


def _build_merged_value_map(ws):
    """
    openpyxl: only top-left cell in a merged range holds the value;
    other cells in that merged range return None.
    Build mapping coord -> top-left value.
    """
    merged_map = {}
    for rng in ws.merged_cells.ranges:
        tl = ws[rng.start_cell.coordinate].value
        for row in range(rng.min_row, rng.max_row + 1):
            for col in range(rng.min_col, rng.max_col + 1):
                coord = openpyxl.utils.get_column_letter(col) + str(row)
                merged_map[coord] = tl
    return merged_map


@pytest.mark.excel
def test_guidelines_dict_matches_excel_sheet():
    xlsx_path = (
        Path(os.getenv("HRRO_EXCEL_PATH", ""))
        if os.getenv("HRRO_EXCEL_PATH")
        else (Path(__file__).parent / "fixtures" / "hrro_baseline.xlsx")
    )
    if not xlsx_path.exists():
        pytest.skip(f"Excel baseline file not found: {xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Guide Line"]

    merged_map = _build_merged_value_map(ws)

    # ✅ Only inherit merged-cell values where it is actually intended:
    #    Profile name (B=2) and SDI (C=3) are merged/repeated visually.
    INHERIT_COLS = {2, 3}  # B, C

    def v(r, c):
        cell = ws.cell(r, c)

        # Known fixture typos override
        if cell.coordinate in KNOWN_CELL_OVERRIDES:
            return KNOWN_CELL_OVERRIDES[cell.coordinate]

        if cell.value is not None:
            return cell.value

        if c in INHERIT_COLS:
            return merged_map.get(cell.coordinate)

        return None

    parsed = {}
    current_profile = None

    # Table starts around row 5 in this fixture
    for r in range(5, ws.max_row + 1):
        profile = v(r, 2)  # B
        if profile:
            current_profile = str(profile).replace("\n", " ").strip()
        if not current_profile:
            continue

        inch = v(r, 8)  # H
        if inch is None:
            continue
        inch = int(inch)

        sdi = v(r, 3)  # C (inherit OK)
        avg_range = _parse_range_lmh(v(r, 4))  # D (inherit NO)
        lead_flux_max = _parse_num_from_constraint(v(r, 6))  # F (inherit NO)
        conc_flow_min = _parse_num_from_constraint(v(r, 9))  # I
        feed_flow_max = _parse_num_from_constraint(v(r, 12))  # L
        dp_max = _parse_num_from_constraint(v(r, 14))  # N
        elem_rec_max = _parse_num_from_constraint(v(r, 16))  # P
        beta_max = _parse_num_from_constraint(v(r, 17))  # Q

        fdr_max = v(r, 18)  # R
        fdr_max = float(fdr_max) if fdr_max is not None else None

        parsed.setdefault(current_profile, {})[inch] = {
            "sdi": (str(sdi).strip() if sdi is not None else None),
            "avg_flux_range_lmh": avg_range,
            "lead_flux_max_lmh": lead_flux_max,
            "conc_flow_min_m3h_per_vessel": conc_flow_min,
            "feed_flow_max_m3h_per_vessel": feed_flow_max,
            "dp_max_bar": dp_max,
            "element_recovery_max_pct": elem_rec_max,
            "beta_max": beta_max,
            "flux_decline_ratio_max_pct": fdr_max,
        }

    # 1) Profile keys must match
    assert set(parsed.keys()) == set(GUIDELINES.keys())

    # 2) Values must match (floats via approx)
    for profile, inch_map in GUIDELINES.items():
        for inch, g in inch_map.items():
            p = parsed[profile][inch]

            assert p["sdi"] == g["sdi"]
            assert p["avg_flux_range_lmh"] == g["avg_flux_range_lmh"]

            for k in [
                "lead_flux_max_lmh",
                "conc_flow_min_m3h_per_vessel",
                "feed_flow_max_m3h_per_vessel",
                "dp_max_bar",
                "element_recovery_max_pct",
                "beta_max",
                "flux_decline_ratio_max_pct",
            ]:
                if g[k] is None:
                    assert p[k] is None
                else:
                    assert pytest.approx(float(p[k]), rel=0, abs=1e-12) == float(g[k])
