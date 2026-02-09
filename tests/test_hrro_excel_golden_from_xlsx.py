import os
from pathlib import Path

import pytest
import openpyxl

from app.services.simulation.modules.hrro import ExcelInputs, compute_excel


def _num(x):
    return float(x) if x is not None else None


@pytest.mark.excel
def test_hrro_compute_excel_matches_xlsx_cells_full():
    xlsx_path = (
        Path(os.getenv("HRRO_EXCEL_PATH", ""))
        if os.getenv("HRRO_EXCEL_PATH")
        else (Path(__file__).parent / "fixtures" / "hrro_baseline.xlsx")
    )
    if not xlsx_path.exists():
        pytest.skip(f"Excel baseline file not found: {xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["CCRO표지"]

    # 캐시값이 없는(=엑셀에서 계산/저장 안 된) 파일이면 여기서 즉시 실패하게
    assert (
        ws["G8"].value is not None
    ), "Excel must be opened/recalculated/saved so formula cached values exist."

    inp = ExcelInputs(
        q_raw_m3h=_num(ws["C7"].value),
        ccro_recovery_pct=_num(ws["G7"].value),
        pf_feed_ratio_pct=_num(ws["K5"].value),
        pf_recovery_pct=_num(ws["K6"].value),
        cc_recycle_m3h_per_pv=_num(ws["O5"].value),
        vessel_count=int(_num(ws["H18"].value)),
        elements_per_vessel=int(_num(ws["H19"].value)),
        area_m2_per_element=_num(ws["C22"].value),
    )

    r = compute_excel(inp)

    # ---- Area / totals ----
    assert pytest.approx(r.total_area_m2, rel=0, abs=1e-9) == _num(ws["H24"].value)

    # ---- CCRO outputs ----
    assert pytest.approx(r.ccro_qp_m3h, rel=0, abs=1e-9) == _num(ws["G8"].value)
    assert pytest.approx(r.ccro_qc_m3h, rel=0, abs=1e-9) == _num(ws["G9"].value)
    assert pytest.approx(r.ccro_flux_lmh, rel=0, abs=1e-9) == _num(ws["G10"].value)

    # ---- PF helper intermediates ----
    assert pytest.approx(r.pf_feed_m3h, rel=0, abs=1e-9) == _num(
        ws["V9"].value
    )  # V9 == pf_feed
    assert pytest.approx(r.cc_feed_m3h, rel=0, abs=1e-9) == _num(ws["V10"].value)  # V10

    # ---- PF outputs ----
    assert pytest.approx(r.pf_feed_m3h, rel=0, abs=1e-9) == _num(ws["K7"].value)
    assert pytest.approx(r.pf_qp_m3h, rel=0, abs=1e-9) == _num(ws["K8"].value)
    assert pytest.approx(r.pf_qc_m3h, rel=0, abs=1e-9) == _num(ws["K9"].value)
    assert pytest.approx(r.pf_flux_lmh, rel=0, abs=1e-9) == _num(ws["K10"].value)

    # ---- CC outputs ----
    assert pytest.approx(r.cc_recovery_pct, rel=0, abs=1e-9) == _num(ws["O6"].value)
    assert pytest.approx(r.cc_blend_feed_m3h_per_pv, rel=0, abs=1e-9) == _num(
        ws["O8"].value
    )
    assert pytest.approx(r.cc_qp_m3h_per_pv, rel=0, abs=1e-9) == _num(ws["O9"].value)
    assert pytest.approx(r.cc_qc_m3h_per_pv, rel=0, abs=1e-9) == _num(ws["O10"].value)
    assert pytest.approx(r.cc_flux_lmh, rel=0, abs=1e-9) == _num(ws["O11"].value)
