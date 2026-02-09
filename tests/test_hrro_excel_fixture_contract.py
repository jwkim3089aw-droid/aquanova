import os
from pathlib import Path

import pytest
import openpyxl


@pytest.mark.excel
def test_hrro_baseline_excel_formulas_are_unchanged():
    xlsx_path = (
        Path(os.getenv("HRRO_EXCEL_PATH", ""))
        if os.getenv("HRRO_EXCEL_PATH")
        else (Path(__file__).parent / "fixtures" / "hrro_baseline.xlsx")
    )
    if not xlsx_path.exists():
        pytest.skip(f"Excel baseline file not found: {xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    ws = wb["CCRO표지"]

    # 핵심 수식 “문자열 계약”
    expected = {
        "G8": "=C7*G7/100",
        "G9": "=C7-G8",
        "G10": "=(G8*1000)/H24",
        "T7": "=IF(K5>=101,((C7*G7%)/10),0)",
        "V10": "=V9/K5%",
        "K7": "=V9",
        "K8": "=K7*K6%",
        "K9": "=K7-K8",
        "K10": "=(K8*1000)/H24",
        "O6": "=(O9/O8)*100",
        "O8": "=O10+O7",
        "O9": "=O7",
        "O10": "=O5",
        "O11": "=(O9*1000)/H24",
    }

    for addr, formula in expected.items():
        assert (
            ws[addr].value == formula
        ), f"{addr} formula changed: {ws[addr].value} != {formula}"
